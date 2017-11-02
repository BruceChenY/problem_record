from openpyxl import load_workbook

wb=load_workbook('test.xlsx')
print(wb.get_sheet_names())
ws=wb['Sheet1']
ws['B4']='12345'
ws['A11']='12345'
wb.save('document.xlsx')
print('OK')